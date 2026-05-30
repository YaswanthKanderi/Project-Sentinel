import os
import sys
from datetime import datetime, timezone
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.makedirs(config.OUTPUT_DIR, exist_ok=True)

HISTORY_FILE = os.path.join(config.OUTPUT_DIR, "extraction_history.xlsx")

def get_or_create_workbook():
    if os.path.exists(HISTORY_FILE):
        return openpyxl.load_workbook(HISTORY_FILE)
    wb = openpyxl.Workbook()
    return wb

def style_header(ws, headers):
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

def export_extraction_history(run_summary):
    wb = get_or_create_workbook()

    # --- Sheet 1: Run History ---
    if "Run History" in wb.sheetnames:
        ws_history = wb["Run History"]
    else:
        ws_history = wb.active
        ws_history.title = "Run History"
        headers = ["Run Timestamp", "Module", "Records Extracted", "Output File", "Status"]
        style_header(ws_history, headers)
        ws_history.column_dimensions["A"].width = 25
        ws_history.column_dimensions["B"].width = 25
        ws_history.column_dimensions["C"].width = 20
        ws_history.column_dimensions["D"].width = 45
        ws_history.column_dimensions["E"].width = 15

    for entry in run_summary:
        ws_history.append([
            entry.get("timestamp"),
            entry.get("module"),
            entry.get("record_count"),
            entry.get("output_file"),
            entry.get("status", "Success")
        ])

    # --- Sheet 2: File Modification History ---
    if "File Modifications" not in wb.sheetnames:
        ws_files = wb.create_sheet("File Modifications")
        file_headers = ["File Name", "Site", "Created By", "Modified By", "Created At", "Modified At", "Size (bytes)", "Version Count", "Sensitivity Label"]
        style_header(ws_files, file_headers)
        ws_files.column_dimensions["A"].width = 25
        ws_files.column_dimensions["B"].width = 25
        ws_files.column_dimensions["C"].width = 30
        ws_files.column_dimensions["D"].width = 30
        ws_files.column_dimensions["E"].width = 25
        ws_files.column_dimensions["F"].width = 25
        ws_files.column_dimensions["G"].width = 15
        ws_files.column_dimensions["H"].width = 15
        ws_files.column_dimensions["I"].width = 20

    # --- Sheet 3: Access & Permission History ---
    if "Access History" not in wb.sheetnames:
        ws_access = wb.create_sheet("Access History")
        access_headers = ["File Name", "Permission ID", "Granted To", "Roles", "Link Scope", "Is Anonymous", "Expiry"]
        style_header(ws_access, access_headers)
        ws_access.column_dimensions["A"].width = 25
        ws_access.column_dimensions["B"].width = 40
        ws_access.column_dimensions["C"].width = 35
        ws_access.column_dimensions["D"].width = 15
        ws_access.column_dimensions["E"].width = 15
        ws_access.column_dimensions["F"].width = 15
        ws_access.column_dimensions["G"].width = 25

    # --- Sheet 4: Audit Sign-In Summary ---
    if "Sign-In Summary" not in wb.sheetnames:
        ws_signin = wb.create_sheet("Sign-In Summary")
        signin_headers = ["Run Timestamp", "Total Records", "Successful Logins", "Failed Logins", "Risky Sign-ins", "High Failure Rate", "Risky Present"]
        style_header(ws_signin, signin_headers)
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws_signin.column_dimensions[col].width = 22

    wb.save(HISTORY_FILE)
    return HISTORY_FILE

def append_file_records(file_records):
    wb = get_or_create_workbook()
    if "File Modifications" not in wb.sheetnames:
        export_extraction_history([])
        wb = get_or_create_workbook()
    ws = wb["File Modifications"]
    for r in file_records:
        ws.append([
            r.get("file_name"),
            r.get("site_name"),
            r.get("created_by_email"),
            r.get("modified_by_email"),
            r.get("created_datetime"),
            r.get("modified_datetime"),
            r.get("size_bytes"),
            r.get("version_count"),
            r.get("sensitivity_label") or "Unlabelled"
        ])
    wb.save(HISTORY_FILE)

def append_permission_records(perm_records):
    wb = get_or_create_workbook()
    if "Access History" not in wb.sheetnames:
        export_extraction_history([])
        wb = get_or_create_workbook()
    ws = wb["Access History"]
    for p in perm_records:
        ws.append([
            p.get("file_name"),
            p.get("permission_id"),
            p.get("granted_to_email"),
            str(p.get("roles", [])),
            p.get("link_scope"),
            p.get("is_anonymous"),
            p.get("expiry")
        ])
    wb.save(HISTORY_FILE)

def append_signin_summary(analysis, timestamp):
    wb = get_or_create_workbook()
    if "Sign-In Summary" not in wb.sheetnames:
        export_extraction_history([])
        wb = get_or_create_workbook()
    ws = wb["Sign-In Summary"]
    flags = analysis.get("governance_flags", {})
    ws.append([
        timestamp,
        analysis.get("total_records"),
        analysis.get("successful_logins"),
        analysis.get("failed_logins"),
        analysis.get("risky_signins"),
        flags.get("high_failure_rate"),
        flags.get("risky_signin_present")
    ])
    wb.save(HISTORY_FILE)